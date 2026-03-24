"""
Tests for export functionality.
"""
import sys
import os
import uuid
import json
import tempfile
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest
from unittest.mock import MagicMock, patch
from datetime import datetime


def make_mock_db():
    """Create a mock DB session that returns empty results for queries."""
    db = MagicMock()

    def mock_query(model):
        q = MagicMock()
        q.filter.return_value = q
        q.all.return_value = []
        q.first.return_value = None
        q.count.return_value = 0
        return q

    db.query.side_effect = mock_query
    return db


def make_audit_run():
    run = MagicMock()
    run.id = uuid.uuid4()
    run.base_url = "https://example.com"
    run.mode = "quick_scan"
    run.status = "completed"
    run.created_at = datetime.utcnow()
    run.started_at = datetime.utcnow()
    run.completed_at = datetime.utcnow()
    run.pages_discovered = 5
    run.pages_crawled = 5
    run.pages_failed = 0
    run.config_id = uuid.uuid4()
    run.error_message = None
    return run


class TestExcelExport:
    def test_excel_file_created(self):
        from worker.exports.excel_exporter import generate_excel

        db = make_mock_db()
        audit_run = make_audit_run()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "test_report.xlsx")
            result_path = generate_excel(db, audit_run, output_path)

            assert os.path.exists(result_path)
            assert result_path == output_path
            assert os.path.getsize(result_path) > 0

    def test_excel_has_expected_sheets(self):
        from worker.exports.excel_exporter import generate_excel
        import openpyxl

        db = make_mock_db()
        audit_run = make_audit_run()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "test_report.xlsx")
            generate_excel(db, audit_run, output_path)

            wb = openpyxl.load_workbook(output_path)
            sheet_names = wb.sheetnames

            expected_sheets = [
                "Summary", "Pages", "Vendors", "Issues",
                "Broken_Requests", "Console_Errors", "Scripts",
                "Cookies_Storage", "Recommendations"
            ]
            for sheet in expected_sheets:
                assert sheet in sheet_names, f"Sheet '{sheet}' missing from Excel workbook"

    def test_excel_summary_sheet_has_audit_data(self):
        from worker.exports.excel_exporter import generate_excel
        import openpyxl

        db = make_mock_db()
        audit_run = make_audit_run()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "test_report.xlsx")
            generate_excel(db, audit_run, output_path)

            wb = openpyxl.load_workbook(output_path)
            ws = wb["Summary"]

            # Check that the base URL appears in the sheet
            found_url = False
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and "example.com" in str(cell):
                        found_url = True
                        break
            assert found_url, "Base URL not found in Summary sheet"


class TestJsonSummaryExport:
    def test_json_file_created(self):
        from worker.exports.report_exporter import generate_json_summary

        db = make_mock_db()
        audit_run = make_audit_run()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "audit_summary.json")
            result_path = generate_json_summary(db, audit_run, output_path)

            assert os.path.exists(result_path)
            assert os.path.getsize(result_path) > 0

    def test_json_has_expected_keys(self):
        from worker.exports.report_exporter import generate_json_summary

        db = make_mock_db()
        audit_run = make_audit_run()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "audit_summary.json")
            generate_json_summary(db, audit_run, output_path)

            with open(output_path, "r") as f:
                data = json.load(f)

            assert "audit" in data
            assert "vendors" in data
            assert "issues" in data
            assert "issue_counts" in data
            assert "total_issues" in data
            assert "generated_at" in data

    def test_json_audit_section_has_correct_values(self):
        from worker.exports.report_exporter import generate_json_summary

        db = make_mock_db()
        audit_run = make_audit_run()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "audit_summary.json")
            generate_json_summary(db, audit_run, output_path)

            with open(output_path, "r") as f:
                data = json.load(f)

            assert data["audit"]["base_url"] == "https://example.com"
            assert data["audit"]["pages_crawled"] == 5
            assert data["audit"]["status"] == "completed"

    def test_json_issue_counts_structure(self):
        from worker.exports.report_exporter import generate_json_summary

        db = make_mock_db()
        audit_run = make_audit_run()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "audit_summary.json")
            generate_json_summary(db, audit_run, output_path)

            with open(output_path, "r") as f:
                data = json.load(f)

            counts = data["issue_counts"]
            assert "critical" in counts
            assert "high" in counts
            assert "medium" in counts
            assert "low" in counts


class TestMarkdownReport:
    def test_markdown_file_created(self):
        from worker.exports.report_exporter import generate_markdown_report

        db = make_mock_db()
        audit_run = make_audit_run()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "audit_summary.md")
            result_path = generate_markdown_report(db, audit_run, output_path)

            assert os.path.exists(result_path)
            assert os.path.getsize(result_path) > 0

    def test_markdown_contains_expected_sections(self):
        from worker.exports.report_exporter import generate_markdown_report

        db = make_mock_db()
        audit_run = make_audit_run()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "audit_summary.md")
            generate_markdown_report(db, audit_run, output_path)

            with open(output_path, "r") as f:
                content = f.read()

            assert "# Odit Tracking Audit Report" in content
            assert "## Summary" in content
            assert "## Detected Vendors" in content
            assert "## Issues" in content
            assert "example.com" in content


class TestHtmlReport:
    def test_html_file_created(self):
        from worker.exports.report_exporter import generate_html_report

        db = make_mock_db()
        audit_run = make_audit_run()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "audit_summary.html")
            result_path = generate_html_report(db, audit_run, output_path)

            assert os.path.exists(result_path)
            assert os.path.getsize(result_path) > 0

    def test_html_is_valid_structure(self):
        from worker.exports.report_exporter import generate_html_report

        db = make_mock_db()
        audit_run = make_audit_run()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "audit_summary.html")
            generate_html_report(db, audit_run, output_path)

            with open(output_path, "r") as f:
                content = f.read()

            assert "<!DOCTYPE html>" in content
            assert "<html" in content
            assert "</html>" in content
            assert "example.com" in content
            assert "Odit" in content
