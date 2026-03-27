"""
Excel exporter: generates a comprehensive Excel workbook using openpyxl.
"""
import os
import logging
from datetime import datetime
from typing import List
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from worker.exports.report_helpers import (
    build_cookie_register,
    build_tag_attribution,
    build_performance_stats,
    compute_executive_metrics,
)

logger = logging.getLogger("odit.worker.excel_exporter")

SEVERITY_COLORS = {
    "critical": "FFCCCC",
    "high":     "FFE0CC",
    "medium":   "FFFACC",
    "low":      "CCE5FF",
}

HEADER_FILL    = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT    = Font(bold=True, color="FFFFFF", size=10)
SUBHEADER_FILL = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)

CATEGORY_COLORS = {
    "analytics":   "EDE9FE",  # purple-100
    "marketing":   "FEE2E2",  # red-100
    "consent":     "D1FAE5",  # green-100
    "functional":  "DBEAFE",  # blue-100
    "security":    "FEF3C7",  # yellow-100
    "unknown":     "F3F4F6",  # gray-100
}


def _write_header_row(ws, headers: List[str]) -> None:
    row = ws.max_row + 1 if ws.max_row > 1 else 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _auto_size_columns(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, max_width)


def generate_excel(db: Session, audit_run, output_path: str) -> str:
    from app.models import (
        PageVisit, NetworkRequest, ConsoleEvent, DetectedVendor, Issue, AuditConfig
    )

    audit_id = audit_run.id
    config = db.query(AuditConfig).filter(AuditConfig.id == audit_run.config_id).first()

    pages    = db.query(PageVisit).filter(PageVisit.audit_run_id == audit_id).all()
    requests = db.query(NetworkRequest).filter(NetworkRequest.audit_run_id == audit_id).all()
    console_events = db.query(ConsoleEvent).filter(ConsoleEvent.audit_run_id == audit_id).all()

    # Audit-level vendors (page_visit_id IS NULL)
    vendors = (
        db.query(DetectedVendor)
        .filter(DetectedVendor.audit_run_id == audit_id, DetectedVendor.page_visit_id == None)
        .all()
    )
    # Page-level vendors (for vendor_id → key mapping in performance)
    page_vendors = (
        db.query(DetectedVendor)
        .filter(DetectedVendor.audit_run_id == audit_id, DetectedVendor.page_visit_id != None)
        .all()
    )

    issues = db.query(Issue).filter(Issue.audit_run_id == audit_id).all()
    issues_sorted = sorted(
        issues,
        key=lambda i: {"critical": 0, "high": 1, "medium": 2, "low": 3}.get(i.severity, 99)
    )

    # Pre-compute derived data
    exec_metrics  = compute_executive_metrics(vendors, issues, config, pages)
    cookie_reg    = build_cookie_register(pages)
    attributions  = build_tag_attribution(vendors, pages, requests)
    perf_stats    = build_performance_stats(vendors, requests, page_vendors)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Executive Summary ────────────────────────────────────────
    ws_exec = wb.create_sheet("Executive Summary")
    ws_exec.column_dimensions["A"].width = 32
    ws_exec.column_dimensions["B"].width = 55

    title_cell = ws_exec.cell(row=1, column=1, value="Odit Tracking Audit — Executive Summary")
    title_cell.font = Font(bold=True, size=14, color="1E3A5F")
    ws_exec.cell(row=2, column=1, value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}").font = Font(italic=True, size=10)

    risk_color_hex = exec_metrics["risk_color"].lstrip("#")
    risk_fill = PatternFill(start_color=risk_color_hex, end_color=risk_color_hex, fill_type="solid")

    exec_data = [
        ("", ""),
        ("RISK OVERVIEW", ""),
        ("Risk Rating",   exec_metrics["risk_rating"]),
        ("Risk Score",    f"{exec_metrics['risk_score']} pts (critical×10, high×5, medium×2, low×1)"),
        ("", ""),
        ("AUDIT SCOPE", ""),
        ("Base URL",      audit_run.base_url),
        ("Mode",          audit_run.mode if isinstance(audit_run.mode, str) else audit_run.mode.value),
        ("Consent Mode",  exec_metrics["consent_mode"]),
        ("Pages Crawled", exec_metrics["pages_crawled"]),
        ("Vendors Found", exec_metrics["vendor_count"]),
        ("Total Issues",  exec_metrics["total_issues"]),
        ("", ""),
        ("ISSUE BREAKDOWN", ""),
        ("Critical",      exec_metrics["issue_counts"]["critical"]),
        ("High",          exec_metrics["issue_counts"]["high"]),
        ("Medium",        exec_metrics["issue_counts"]["medium"]),
        ("Low",           exec_metrics["issue_counts"]["low"]),
        ("", ""),
        ("COMPLIANCE POSTURE", ""),
        ("GDPR / ePrivacy", exec_metrics["gdpr_posture"]),
        ("CCPA / CPRA",    exec_metrics["ccpa_posture"]),
        ("", ""),
        ("TOP FINDINGS", ""),
    ]

    for i, (key, val) in enumerate(exec_data, start=4):
        key_cell = ws_exec.cell(row=i, column=1, value=key)
        val_cell = ws_exec.cell(row=i, column=2, value=val)
        if key in ("RISK OVERVIEW", "AUDIT SCOPE", "ISSUE BREAKDOWN", "COMPLIANCE POSTURE", "TOP FINDINGS"):
            key_cell.font = Font(bold=True, size=10, color="1E3A5F")
            key_cell.fill = SUBHEADER_FILL
            val_cell.fill = SUBHEADER_FILL
        else:
            key_cell.font = Font(bold=True, size=10)
        if key == "Risk Rating" and val:
            val_cell.font = Font(bold=True, color="FFFFFF")
            val_cell.fill = risk_fill

    # Top findings rows
    top_row = len(exec_data) + 4
    for issue in exec_metrics["top_issues"]:
        ws_exec.cell(row=top_row, column=1, value=f"[{issue.severity.upper()}]").font = Font(bold=True)
        ws_exec.cell(row=top_row, column=2, value=issue.title)
        top_row += 1

    # ── Sheet 2: Summary (stats) ──────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 50

    summary_data = [
        ("Audit ID",          str(audit_run.id)),
        ("Base URL",          audit_run.base_url),
        ("Mode",              audit_run.mode if isinstance(audit_run.mode, str) else audit_run.mode.value),
        ("Status",            audit_run.status if isinstance(audit_run.status, str) else audit_run.status.value),
        ("Created At",        str(audit_run.created_at)),
        ("Started At",        str(audit_run.started_at) if audit_run.started_at else "—"),
        ("Completed At",      str(audit_run.completed_at) if audit_run.completed_at else "—"),
        ("", ""),
        ("Pages Discovered",  audit_run.pages_discovered),
        ("Pages Crawled",     audit_run.pages_crawled),
        ("Pages Failed",      audit_run.pages_failed),
        ("", ""),
        ("Vendors Detected",  len(vendors)),
        ("Total Issues",      len(issues)),
        ("Critical Issues",   exec_metrics["issue_counts"]["critical"]),
        ("High Issues",       exec_metrics["issue_counts"]["high"]),
        ("Medium Issues",     exec_metrics["issue_counts"]["medium"]),
        ("Low Issues",        exec_metrics["issue_counts"]["low"]),
        ("", ""),
        ("Risk Rating",       exec_metrics["risk_rating"]),
        ("GDPR Posture",      exec_metrics["gdpr_posture"]),
        ("CCPA Posture",      exec_metrics["ccpa_posture"]),
    ]

    ws_summary.cell(row=1, column=1, value="Odit Tracking Audit Summary").font = Font(bold=True, size=14)
    ws_summary.cell(row=2, column=1, value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}").font = Font(italic=True, size=10)

    for i, (key, val) in enumerate(summary_data, start=4):
        ws_summary.cell(row=i, column=1, value=key).font = Font(bold=True)
        ws_summary.cell(row=i, column=2, value=val)

    # ── Sheet 3: Pages ────────────────────────────────────────────────────
    ws_pages = wb.create_sheet("Pages")
    _write_header_row(ws_pages, [
        "URL", "Final URL", "Status Code", "Title", "Group",
        "Load Time (ms)", "Console Errors", "Failed Requests",
        "Vendor Count", "Crawled At", "Error"
    ])

    from collections import defaultdict
    page_vendor_counts = defaultdict(int)
    for pv in page_vendors:
        page_vendor_counts[str(pv.page_visit_id)] += 1

    for page in pages:
        ws_pages.append([
            page.url,
            page.final_url,
            page.status_code,
            page.page_title,
            page.page_group,
            round(page.load_time_ms, 1) if page.load_time_ms else None,
            page.console_error_count,
            page.failed_request_count,
            page_vendor_counts.get(str(page.id), 0),
            str(page.crawled_at) if page.crawled_at else None,
            page.error_message,
        ])
    _auto_size_columns(ws_pages)

    # ── Sheet 4: Vendors ──────────────────────────────────────────────────
    ws_vendors = wb.create_sheet("Vendors")
    _write_header_row(ws_vendors, [
        "Vendor Key", "Vendor Name", "Category", "Detection Method",
        "Load Source", "Page Count", "Evidence"
    ])
    for v in vendors:
        load_src = attributions.get(v.vendor_key, "Unknown")
        ws_vendors.append([
            v.vendor_key,
            v.vendor_name,
            v.category,
            v.detection_method,
            load_src,
            v.page_count,
            str(v.evidence) if v.evidence else "",
        ])
    _auto_size_columns(ws_vendors)

    # ── Sheet 5: Issues ───────────────────────────────────────────────────
    ws_issues = wb.create_sheet("Issues")
    _write_header_row(ws_issues, [
        "Severity", "Category", "Title", "Description",
        "Affected URL", "Affected Vendor", "Likely Cause",
        "Recommendation", "Remediation Steps"
    ])
    for issue in issues_sorted:
        row_idx = ws_issues.max_row + 1
        ws_issues.append([
            issue.severity,
            issue.category,
            issue.title,
            issue.description,
            issue.affected_url or "",
            issue.affected_vendor_key or "",
            issue.likely_cause or "",
            issue.recommendation or "",
            issue.remediation_steps or "",
        ])
        if issue.severity in SEVERITY_COLORS:
            fill = PatternFill(
                start_color=SEVERITY_COLORS[issue.severity],
                end_color=SEVERITY_COLORS[issue.severity],
                fill_type="solid"
            )
            for col in range(1, 10):
                ws_issues.cell(row=row_idx, column=col).fill = fill
    _auto_size_columns(ws_issues)

    # ── Sheet 6: Cookie Register ──────────────────────────────────────────
    ws_cookies = wb.create_sheet("Cookie Register")
    _write_header_row(ws_cookies, [
        "Cookie Name", "Domain", "1st / 3rd Party", "Expiry",
        "Purpose Category", "Vendor / Tool", "Description",
        "HttpOnly", "Secure", "SameSite", "Pages Seen"
    ])
    for cookie in cookie_reg:
        row_idx = ws_cookies.max_row + 1
        ws_cookies.append([
            cookie["name"],
            cookie["domain"],
            cookie["party"],
            cookie["expiry"],
            cookie["purpose_category"],
            cookie["vendor"],
            cookie["description"],
            "Yes" if cookie["http_only"] else ("No" if cookie["http_only"] is not None else ""),
            "Yes" if cookie["secure"] else ("No" if cookie["secure"] is not None else ""),
            cookie.get("same_site", ""),
            cookie["page_count"],
        ])
        # Color-code by purpose category
        cat = cookie["purpose_category"]
        if cat in CATEGORY_COLORS:
            fill = PatternFill(
                start_color=CATEGORY_COLORS[cat],
                end_color=CATEGORY_COLORS[cat],
                fill_type="solid"
            )
            for col in range(1, 12):
                ws_cookies.cell(row=row_idx, column=col).fill = fill
    _auto_size_columns(ws_cookies)

    # ── Sheet 7: Performance ──────────────────────────────────────────────
    ws_perf = wb.create_sheet("Performance")
    _write_header_row(ws_perf, [
        "Vendor", "Category", "Requests Captured",
        "Avg Load Time (ms)", "Max Load Time (ms)", "Pages Affected"
    ])
    for stat in perf_stats:
        ws_perf.append([
            stat["vendor_name"],
            stat["category"],
            stat["request_count"],
            stat["avg_timing_ms"],
            stat["max_timing_ms"],
            stat["pages_affected"],
        ])
    _auto_size_columns(ws_perf)

    # ── Sheet 8: Broken_Requests ──────────────────────────────────────────
    ws_broken = wb.create_sheet("Broken_Requests")
    _write_header_row(ws_broken, [
        "URL", "Method", "Status Code", "Resource Type",
        "Is Tracking", "Failed", "Failure Reason", "Page Visit ID"
    ])
    broken_reqs = [r for r in requests if r.failed or (r.status_code and r.status_code >= 400)]
    for req in broken_reqs[:500]:
        ws_broken.append([
            req.url[:200],
            req.method,
            req.status_code,
            req.resource_type,
            req.is_tracking_related,
            req.failed,
            req.failure_reason or "",
            str(req.page_visit_id),
        ])
    _auto_size_columns(ws_broken)

    # ── Sheet 9: Console_Errors ───────────────────────────────────────────
    ws_console = wb.create_sheet("Console_Errors")
    _write_header_row(ws_console, [
        "Level", "Message", "Source URL", "Line", "Column", "Page Visit ID", "Captured At"
    ])
    error_events = [e for e in console_events if e.level == "error"]
    for evt in error_events[:500]:
        ws_console.append([
            evt.level,
            evt.message[:300],
            evt.source_url or "",
            evt.line_number,
            evt.column_number,
            str(evt.page_visit_id),
            str(evt.captured_at),
        ])
    _auto_size_columns(ws_console)

    # ── Sheet 10: Scripts ─────────────────────────────────────────────────
    ws_scripts = wb.create_sheet("Scripts")
    _write_header_row(ws_scripts, ["Page URL", "Script Src", "Is Tracking"])
    script_entries = []
    from worker.detectors.vendor_detector import is_tracking_url
    for page in pages:
        for src in (page.script_srcs or []):
            script_entries.append([page.url, src, is_tracking_url(src)])
    for entry in script_entries[:1000]:
        ws_scripts.append(entry)
    _auto_size_columns(ws_scripts, max_width=80)

    # ── Sheet 11: Recommendations ─────────────────────────────────────────
    ws_recs = wb.create_sheet("Recommendations")
    _write_header_row(ws_recs, [
        "Priority", "Severity", "Issue Title",
        "Recommendation", "Remediation Steps", "Affected Vendor"
    ])
    for i, issue in enumerate(issues_sorted, 1):
        ws_recs.append([
            i,
            issue.severity.upper(),
            issue.title,
            issue.recommendation or "See issue details.",
            issue.remediation_steps or "",
            issue.affected_vendor_key or "",
        ])
    _auto_size_columns(ws_recs)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    logger.info(f"Excel workbook saved: {output_path}")
    return output_path
